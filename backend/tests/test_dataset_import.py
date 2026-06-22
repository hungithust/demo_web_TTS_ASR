import openpyxl
from openpyxl.styles import Font

from app.services.dataset_import import parse_dataset_xlsx


def _make_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [
        (1, "Câu cơ bản", True),   # category header (bold col-B)
        (None, "Xin chào.", False),
        (None, "Hôm nay trời đẹp.", False),
        (2, "Số nguyên", True),    # next category
        (None, "Một hai ba.", False),
    ]
    for i, (a, b, bold) in enumerate(rows, start=1):
        if a is not None:
            ws.cell(row=i, column=1, value=a)
        cell = ws.cell(row=i, column=2, value=b)
        cell.font = Font(bold=bold)
    path = tmp_path / "ds.xlsx"
    wb.save(path)
    return str(path)


def test_parse_groups_texts_under_bold_headers(tmp_path):
    pairs = parse_dataset_xlsx(_make_xlsx(tmp_path))
    assert pairs == [
        ("Câu cơ bản", "Xin chào."),
        ("Câu cơ bản", "Hôm nay trời đẹp."),
        ("Số nguyên", "Một hai ba."),
    ]


def test_parse_skips_blank_cells(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="Cat A").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=2, column=2, value="   ")          # whitespace only -> skip
    ws.cell(row=3, column=2, value=None)            # empty -> skip
    ws.cell(row=4, column=2, value="Real text.")
    path = tmp_path / "b.xlsx"
    wb.save(path)
    assert parse_dataset_xlsx(str(path)) == [("Cat A", "Real text.")]
