import openpyxl


def parse_dataset_xlsx(path: str) -> list[tuple[str, str]]:
    """Parse the TTS dataset xlsx into (category, text) pairs.

    Rule (column B): a non-empty BOLD cell starts a new category; following
    non-bold non-empty cells are that category's text test cases.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    pairs: list[tuple[str, str]] = []
    current: str | None = None
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        value = cell.value
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        if cell.font is not None and cell.font.bold:
            current = text
        elif current is not None:
            pairs.append((current, text))
    return pairs
